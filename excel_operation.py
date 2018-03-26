'''建立一个新的SHEET，将后页SHEET，筛选复制到新建SHEET。
熟悉openpyxl模块的使用。'''
from openpyxl import Workbook,load_workbook
outwb=Workbook()
careerSheet=outwb.create_sheet(0,'career')
for sheetName in inwb.get_sheet_names():
  if not sheetName.isdight():
    continue
  sheet=inwb[sheetNmae]
  
